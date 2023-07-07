from openpyxl.styles import Color, Font, PatternFill, colors
from openpyxl.styles import Font, Color, colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 엑셀 파일 불러오기
workbook = load_workbook(filename='파일이름.xlsx')
worksheet = workbook.active

# 칼럼1과 칼럼2를 저장할 리스트와 시트에서 데이터를 불러와서 저장할 리스트 생성
col1_values = []
col2_values = []
data = []

# 시트에서 데이터를 불러와서 저장
for row in worksheet.iter_rows(min_row=2, values_only=True):
    col1_value = str(row[4])  # 5번째 칸에서 칼럼1 데이터를 문자열로 변환하여 가져옴
    col2_value = str(row[6])  # 7번째 칸에서 칼럼2 데이터를 문자열로 변환하여 가져옴
    other_columns_data = row[:4] + row[5:6] + row[7:]  # 날짜와 관련 없는 칼럼들의 데이터
    col1_values.append(col1_value)
    col2_values.append(col2_value)
    data.append((col1_value, col2_value, *other_columns_data))

# 칼럼1, 칼럼2를 기준으로 데이터 정렬
sorted_data = [x for _, x in sorted(zip(
    zip(col1_values, col2_values), data), key=lambda pair: (pair[0][0], pair[0][1]))]

# 시트에서 데이터 삭제
start_col = column_index_from_string('A')
end_col = column_index_from_string('L')
worksheet.delete_rows(2, worksheet.max_row)

# 정렬된 데이터 시트에 삽입
for i, row_data in enumerate(sorted_data):
    col1, col2, *other_columns_data = row_data
    row = other_columns_data[:4] + [col1] + \
        other_columns_data[4:5] + [col2] + other_columns_data[5:]
    for j, val in enumerate(row):
        worksheet.cell(row=i+2, column=j+start_col, value=val)

# 엑셀 파일 저장
workbook.save(filename='파일이름1.xlsx')

### 이제 폰트와 선을 설정하자!!!!!!!!


filename = '파일이름1.xlsx'
sheetname = 'Sheet1'
wb = load_workbook(filename)
ws = wb[sheetname]

data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(row)

col1_idx = 4  # 칼럼1 인덱스 (0부터 시작)
col2_idx = 6  # 칼럼2 인덱스 (0부터 시작)

# 글자체와 크기 설정
font2 = Font(name='NGULIM', size=12, bold=True)
font1 = Font(name='새굴림', size=11)

# 전체 범위에 테두리 설정
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
max_row = ws.max_row
max_col = ws.max_column
for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
    for cell in row:
        cell.border = border

# 각 칼럼별 글자체와 크기 설정
for col_num in range(1, max_col + 1):
    col_letter = get_column_letter(col_num)
    if col_num in [col1_idx+1, col2_idx+1]:  # 칼럼1과 칼럼2 다음 칼럼
        for cell in ws[col_letter]:
            cell.font = font2
    elif col_num in [col1_idx+1, col2_idx+1]:  # 칼럼1과 칼럼2
        for cell in ws[col_letter]:
            cell.font = font1
    else:
        for cell in ws[col_letter]:
            cell.font = font1

sorted_data = sorted(data, key=lambda x: (x[col1_idx], x[col2_idx]))

for row_num, row in enumerate(sorted_data, start=2):
    for col_num, value in enumerate(row, start=1):
        col_letter = get_column_letter(col_num)
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.border = border
        if col_num in [col1_idx+1, col2_idx+1]:  # 칼럼1과 칼럼2 다음 칼럼
            cell.font = font2
        elif col_num in [col1_idx+1, col2_idx+1]:  # 칼럼1과 칼럼2
            cell.font = font1

wb.save(filename)



filename = '파일이름1.xlsx'
sheetname = 'Sheet1'
wb = load_workbook(filename)
ws = wb[sheetname]

data = []

# 노란색으로 칠하기 위한 PatternFill 객체 생성
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


for i, row in enumerate(ws.iter_rows(min_row=2)):
    if row[2].value == '222':  # 3번째 열(column)의 값이 '222'인 경우
        for cell in row:
            
            cell.font = Font(color=colors.YELLOW)
    else:
        for cell in row:
            
            cell.font = Font(color=colors.YELLOW)
        
            
wb.save('파일이름2.xlsx')
