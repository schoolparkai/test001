import codecs
import sys
import time
import os
import openpyxl, time
wb = openpyxl.load_workbook(
    "C:\\Users\\park9\\OneDrive\\바탕 화면\\시조_수사_전승\\sizo_scan\\시조_DIC.xlsx")

#wb.active
#print(wb.sheetnames)   


sheet = wb['Sheet1']  # 시트 가지고 오기  # 읽을 때부터 utf-16으로 읽어야 함-> 아니면 글자가 깨짐.
# 작품 번호를 가지고 접근해야 함.-> 작품번호로 매치 준비 중임.

#딕셔너리 사전 제작
sizo="교본역대시조전서"
work_num = 99999
copy_num = "이본수"
book = "문헌명"
book_year = "문헌연대"
genre = "장르"
writer = "작가명"
writer_year = "작가연대"
line01 = "초장"
line02 = "중장"
line02 = "종장"

malmungchi={}
malmungchi = {sizo: {work_num: {
    '이본수': 'copy_num', '문헌명': 'book', '문헌연대':'book_year', '장르':'genre', '작가명':'writer', '작가연대':'writer_year', '초장':'line01', '중장':'line02', '종장':'line03' }}}

print("<<딕셔너리를 만들기 시작합니다>>")
print("\n")

    
j=0    
# row(가로) 값을 읽기        #for rows in sheet.iter_rows():
for rows in sheet.iter_rows():
    i=0
    j+=1     
    for cell in rows:            
        i+=1
        if i==1:
            work_num = cell.value
            #work_num = str(work_num1)
        elif i==2:
            copy_num = cell.value            
        elif i==3:
            pass
        elif i==4:
            book = cell.value
        elif i==5:
            book_year = cell.value
        elif i==6:
            genre = cell.value
        elif i==7:
            writer= cell.value
        elif i==8:
            writer_year= cell.value
        elif i==9:
            line01= cell.value
        elif i==10:
            line02= cell.value
        elif i == 11:
            line03 = cell.value
        else:
            pass
        
    
    #'이본수': 'copy_num', '문헌명': 'book', '문헌연대':'book_year', '장르':'genre', 
    #'작가명':'writer', '작가연대':'writer_year', '초장':'line01', '중장':'line02', '종장':'line03' 
 
    #print(researcher_title_name, title, author, text)
    if j>1:
        malmungchi[sizo][work_num] = {}
        malmungchi[sizo][work_num]['이본수'] = copy_num
        malmungchi[sizo][work_num]['문헌명'] =book
        malmungchi[sizo][work_num]['문헌연대'] = book_year
        malmungchi[sizo][work_num]['장르'] = genre
        malmungchi[sizo][work_num]['작가명'] = writer
        malmungchi[sizo][work_num]['작가연대'] = writer_year
        malmungchi[sizo][work_num]['초장'] = line01
        malmungchi[sizo][work_num]['중장'] = line02
        malmungchi[sizo][work_num]['종장'] = line03

    else:
        pass
    
    
    

    print("작품 "+ str(j-1) + "개를 처리하였습니다.")

        
print("\n")
print("*"*100) 
print("작업을 완료하였습니다. 완성된 malmungchi 딕셔너리는 다음과 같습니다.")   
print("\n")
#print(malmungchi)
        
# Dictionary

with open('C:\\Users\\park9\\OneDrive\\바탕 화면\\시조_수사_전승\\sizo_scan\\Sizo_Dic.txt', mode='w', encoding='utf-16') as f:
    f.write(str(malmungchi))

