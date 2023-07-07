#!/usr/bin/python 
# -*- coding: utf-8 -*-
# 검색 결과(텍스트 파일 폴더)를 엑셀로 변환하여 저장
## 작품번호를 list로 받아, 딕셔너리를 불러 필요한 정보를 가져옴 
## 작품번호, 제목, 작자, 창작연대, 검색된 행 등을 Excel로 저장하여 제공토록 함
# 추출함수로 만들어 코드를 간략하게 함.


import os, re
import glob, time
import sys, codecs
import 말뭉치_함수RE
import Sizo_Dic
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


global fff, words, i, count_item, grab, startend, howmany, grabb, each_line1

# ㄷㅏㄱ?ㅏ> .{3,25}ㄴ.ㄴㄷㅡㅅ" # 성공 "ㄴᆞㄴ\s\S.{1,15}ㅇㅏ\s"    /ㅇㅔㅅㅕ.{1,9}ㄹ\s?ㅅㅗㄴㅑ/
words = r"\b\w*?(\w{3,})\w*?\b\W*?\b\w*?(\w{3,})\w*?\b\W*?\w*?\W*?" + "     " + \
    r"\W*?\w*?\b\w*?\1\w*?\b\W*?\b\w*?\2\w*?\b\W*?"
#333 - 성공    
words = r"\b\w*?(\w{3,3})\w*?\b.{1,30}?\b\w*?(\w{3,3})\w*?\b.{1,30}?\b\w*?(\w{3,3})\w*?\b.{1,30}?" + \
    "     " + \
        r".{1,30}?\b\w*?\1\w*?\b.{1,30}?\b\w*?\2\w*?\b.{1,30}?\b\w*?\3\w*?\b"
#2222 -     
words = r"\b\w*?(\w{3,3})\w*?\b.{1,30}?\b\w*?(\w{2,2})\w*?\b.{1,30}?\b\w*?(\w{2,2})\w*?\b.{1,100}?" + "     " +  r".{1,30}?\b\w*?\1\w*?\b.{1,30}?\b\w*?\2\w*?\b.{1,30}?\b\w*?\3\w*?\b"


#66 - 성공    
words = r"\b\w*?(\w{9,})\w*?\b.{1,30}?\b\w*?\1\w*?\b" 
belong = "章內_단어경계外_중첩_9_9_"
words_han ="_동일어휘반복_그룹간거리(10)"
#장외_2장간_교차_5_5 :r"\b\w*?(\w{5,})\w*?\b.{1,30}?\b\w*?(\w{5,})\w*?\b.{1,100}?" + "     " +  r".{1,30}?\b\w*?\2\w*?\b.{1,30}?\b\w*?\1\w*?\b"
# 장외_2장간_2_2_2_2_ : r"\b\w*?(\w{2,2})\w*?\b.{1,30}?\b\w*?(\w{2,2})\w*?\b.{1,30}?\b\w*?(\w{2,2})\w*?\b.{1,30}?.{1,30}?\b\w*?(\w{2,2})\w*?\b.{1,30}?" + "     " +  r".{1,30}?\b\w*?\1\w*?\b.{1,30}?\b\w*?\2\w*?\b.{1,30}?\b\w*?\3\w*?\b.{1,30}?\b\w*?\4\w*?\b"
#(성공:초중 동일 구조의 2회 반복) words = r"\b\w*?(\w{3,})\w*?\b\W*?\b\w*?(\w{3,})\w*?\b.{1,100}" + "     " +  r".{1,100}?\w*?\b\w*?\1\w*?\b\W*?\b\w*?\2\w*?\b\W*?"
# 초중_어절내에서 2회  r"\b\w*?(\w{3,})\w*?\1\w*?\b\W*?\w*?" + "     " + r"\W*?\w*?\b\w*?(\w{3,})\w*?\2\w*?\b"   /에러 있음. 맞는 거 같은데 검색이 안 됨.
# 초중2회 r"\b\w*?(\w{3,})\w*?\b\W*?\w*?\b\w*?\1\w*?\b\W*?\w*?" + "     " +   r"\w*?\b\w*?(\w{3,})\w*?\b\W*?\w*?\b\w*?\2\w*?\b\W*?" 
#특정 어휘가 반복되는 것을 검출 r"\b\w*(ㄷㅗㄹㅏ)\w*\s+\w*\1\w*\b"
# 반복되는 임의의 문자열의 길이를 3글자 이상으로 수정하고, 어절의 앞부분이나 뒷부분일 수 있음 r"\b\w*?(\w{3,})\w*?\b\W*?\b\w*?\1\w*?\b"
# 연속되는 7개의 문자가 반복/ 어미 미검출  r"\b(\S{7,})(\s+\1)+\b"
# 동일 어휘의 3회 반복  r'\b(\w{3,})\b\s+\b\1\b'
# 어절 앞 부분을 기준으로 2회 이상 반복/ 어미 반복 검출못함  r'\b(\w*\w{4,}\w*)(\w+\s+\1)+'
# 정규표현식 문법 ?(앞 문자가 0개 또는 1개)/ .(최소 한개의 문자/공백포함)/ {n,m}(n번 이상 m번 이하 반복: dot와 조합시켜야)/ \s(스페이스) / \S(스페이스 아닌 거) / (그룹){반복횟수}
# 한자로 된 작품은 한음절 크기가 1임에 주의할 것. {범위} 입력 시 1부터 입력할 것.
# \s -> 그냥 띄어쓰기도 됨  #아래아 + ㅣ = 한글자(ᆡ) / .{1,2} 이런 형태로 검색해야 함.
# 띄어쓰기 유무에 주의 -> 선택적으로 검색 : \s?
### 의 다르니  ㅇㅡㅣ ㄴᆡㄷㅏㄹㅡㄴㅣ 
# 연속된 같은 문자 검색 ([a-z])\1
# "(^ㅇㅣ.{1,15}ㄴ\s.{2,40}\sㅈㅕ.{1,15}ㄴ\s)|(\sㅇㅣ.{1,15}ㄴ\s.{2,40}\sㅈㅕ.{1,15}ㄴ\s)" # 성공(그룹으로 선택)
# "\S.{4,11}ㅇㅣ\s\S.{4,11}ㄱㅗ\S.{4,11}ㅇㅣ\s\S.{4,11}ᆡ" # 성공
# "ㄴᆞㄴ\s\S{0,1}ㅈㅓ\s\S.{1,8}ㅇㅏ\s" # 성공
#"ㄴᆞㄴ\s\S{0,1}[ㅈㅓ |ㄷㅕ |ㅈㅕ ]\s\S.{1,8}[ㅇㅏ |ㅇㅑ ]\s" # 성공
# "冬+.{0,6}ㅅ+.{2,4}[ㄷ|ㅼ]+.{1,1}ㄹ"     |ㄷㅕ |ㅈㅕ  |ㅇㅣㅇㅕ |ㅇㅑ 
# "ㄴᆞㄴ\s\S{1,5}ㄹ\s\S.{1,15}ㄴᆞㄴ\s\S{1,5}ㄹ\s" # 는 ㄹ 는 ㄹ #성공

mal="sizo_work"  
#"_data_history"  
#fff=""
fff=[]
sub_fff=[]
startend=[]
grab=1  # 홀수이면 선행 행의 값을 반영하지 않음 
howmany=0
grabb=0
each_line1=""

def sub_fff_TO_fff(count_number,w_num, 비교항 ,author1,author_year, book1, book1_year,반복항, line1, line2, line3, copynum, fff_list):
    count_number+=1
    sub_fff=[]
    sub_fff.append(count_number)
    sub_fff.append(w_num)
    sub_fff.append(비교항)
    #sub_fff.append(genre1)
    sub_fff.append(author1)
    sub_fff.append(author_year)
    sub_fff.append(book1)
    sub_fff.append(book1_year)
    sub_fff.append(반복항)
    #sub_fff.append(each_line)
    sub_fff.append(line1)
    sub_fff.append(line2)
    sub_fff.append(line3)
    sub_fff.append(copynum)
    fff_list.append(sub_fff)
    return count_number, fff_list

                                    
#print(glob.glob("*")) -> glob의 문제가 아님

#path="gasa_work_paired" #아래처럼 절대경로로 바꾸어주니 작동함
#path = "C:\시조_수사_전승\sizo_scan\sizo_work_paired"
path="C:\\Users\\park9\\OneDrive\\바탕 화면\\시조_수사_전승\\sizo_scan\\sizo_work_paired"
data_files = glob.glob(f"{path}/*") #

count_item=0

errors = ""

try:
    for e_f in data_files:
        i=0
        
        with codecs.open(e_f,"r", encoding="utf-16") as f:
            name=f.readline()
            name=f.readline()
            line01_divided=f.readline()
            line01_original=f.readline()
            line02_divided=f.readline()
            line02_original=f.readline()
            line03_divided=f.readline()
            line03_original=f.readline()
            #line0102 = line01_divided + "     " + line02_divided
            #line0203 = line02_divided + "     " + line03_divided
            #line0103 = line01_divided + "     " + line03_divided
            #line010203 = line01_divided + "     " + line02_divided + "     " + line03_divided
            
            #f_lines = [line0102, line0203, line0103, line010203,line03_original]
            f_lines = [line01_divided, line02_divided, line03_divided,line03_original]
            
            #print(name0)
            print(name)
            #time.sleep(1)
            
            
            try:
                #work_num, work_name = name.split(".")
                work_num = int(name[1:])
                work_name = "00000"
            except Exception as ex: # 에러 종류
                print("에러가 발생했습니다.", ex)
                errors = errors + name +"\n"
                #print(name)
 
    
            work_num = int(work_num)
            
            each_line_count = 0
            
            try:
                if len(line02_original)> 26:  #중장이 긴 것은 패스
                    pass
                else:
                    
                    for each_line in f_lines:
                        pattern_search = re.compile(words)
                        c = pattern_search.findall(each_line)  # 검색 건수를 아래에서 howmany로 저장할 것.

                        each_line_count +=1            
                        
                        if len(c)==0 : #검색 값이 없을 때
                            pass
                        elif len(c)>0 : #자모열에서 검색 값이 있을 때,
                            print('work_num :' + str(work_num))
                            반복항변수 =""
                            
                            for i in c:
                                if 반복항변수 == "":
                                    반복항변수 = str(i)
                                else:
                                    반복항변수 = 반복항변수 + "/" + str(i)
                            
                            if each_line_count == 1:
                                비교항변수 = "초01"
                            elif each_line_count == 2:
                                비교항변수 = "중02"
                            elif each_line_count == 3:
                                비교항변수 = "종03"
                            #elif each_line_count == 4:
                            #    비교항변수 = "초01-중02-종03"
                            else:
                                pass
                            #each_line_chosen = each_line1[:]
                                                    
                                        
                            # 문서의 처음/Sizo_Dic과 연동하여 '작가명/창작연대' 변수로 가져올 것
                            #딕셔너리와 연동 -> 작가/창작연대 추출 -> 아래에 이 변수를 반복 활용할 것.
                                                                        
                            copy_num = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['이본수']
                            book = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['문헌명']
                            book_year = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['문헌연대']
                            #genre = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['장르']
                            author = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['작가명']
                            author_year = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['작가연대']                            
                            line01 = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['초장']
                            line02 = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['중장']
                            line03 = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['종장']


                            count_item, fff = sub_fff_TO_fff(count_item,work_num,비교항변수,author,author_year, book, book_year, 반복항변수, line01, line02, line03,copy_num, fff) 
                                
                            
                        
                            print('len(each_line1)                        :' +str(len(each_line1))) #확인용>>>>>>>>>>>>>>>>>>

                            
                                
                        elif each_line==f.readline(-1):
                            f.close()
                            print("f.close()니다")
                            
                        else:                        
                            pass
                                            
            except IOError as ioerr:
                print('File error (outer try): ' + str(ioerr))
    
                
    if not len(fff)==0 :
        #out=open('C:\\가사작업중20200208\\Gasa_Divide\\result\\result_'+'('+str(count_item)+'건)'+words_han+"_from_"+mal+'.txt', mode= 'w', encoding='utf-16')
        #print(fff, file=out)
        
        #엑셀로 저장
        wb = Workbook()
        ws = wb.active
        ws.title = words_han
        thin_border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
        
        x = 1
        y = 0 
        for s_fff in fff:
            #제목 행 세팅
            ws.cell(column = 1, row = 1, value = "연번").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 1, row = 1, value = "연번").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 2, row = 1, value = "작품번호").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 2, row = 1, value = "작품번호").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 3, row = 1, value = "검색된 위치").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 3, row = 1, value = "검색된 위치").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 4, row = 1, value = "작가").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 4, row = 1, value = "작가").font = Font(name="맑은 고딕", size=11, color="00FF0000")
            ws.cell(column = 5, row = 1, value = "작가연대").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 5, row = 1, value = "작가연대").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 6, row = 1, value = "가집").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 6, row = 1, value = "가집").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 7, row = 1, value = "가집연대").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 7, row = 1, value = "가집연대").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 8, row = 1, value = "반복항").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 8, row = 1, value = "반복항").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 9, row = 1, value = "초장").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 9, row = 1, value = "초장").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 10, row = 1, value = "중장").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 10, row = 1, value = "중장").font = Font(name="맑은 고딕", size=11, color="00FF0000")
            ws.cell(column = 11, row = 1, value = "종장").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 11, row = 1, value = "종장").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 12, row = 1, value = "이본수").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 12, row = 1, value = "이본수").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            


       

            # 2열부터 검색된 내용 정리
            x+=1
            y=0
            for item in s_fff:
                y+=1
                #ws.cell = Font(name ="NGULIM", size = 20)
                cell_address = ws.cell(column = y, row = x, value = item)
                #print(cell_address.coordinate)
                cell_address_obj = ws[cell_address.coordinate]
                cell_address_obj.font = Font(name ="새굴림", size = 10)
                
                #ws[aaaa].font  = Font(name ="NGULIM", size = 20)
                #ws.cell = Font(name ="NGULIM", size = 20)
                cell_address_obj.border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
        
        # 열의 넓이 정하기
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 7
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 29
        ws.column_dimensions['I'].width = 60
        ws.column_dimensions['J'].width = 60
        ws.column_dimensions['K'].width = 60
        ws.column_dimensions['L'].width = 6
                
        wb.save('C:\\Users\\park9\\OneDrive\\바탕 화면\\시조_수사_전승\\sizo_scan\\result_pattern\\' + belong +'result_' +
                '('+str(count_item)+'건)'+words_han+"_from_"+mal+'.xlsx')
        wb.close
       
        
        #ffff="error checking"   # 저장 경로 확인해 보니 c:\사용자\park9 
        #with open('resul9999t.txt', mode= 'w', encoding= 'utf-16') as f:
        #    f.write(fff)
                  
        #out.close()
        print("검색어는 '" +str(words)+"' 입니다")
        
        print("검색 건수는 모두 '" +str(count_item)+"'건 입니다. 엑셀 파일로 저장하였습니다. ")
        
        
    else:
        print("해당되는 자료가 없습니다")
        
                    
except IOError as ioerr:
    print('File error (outer try): ' + str(ioerr))

print(errors)
    


    
    
    
