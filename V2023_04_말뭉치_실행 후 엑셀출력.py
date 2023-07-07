#!/usr/bin/python 
# -*- coding: utf-8 -*-
# 검색 결과(텍스트 파일 폴더)를 엑셀로 변환하여 저장
## 작품번호를 list로 받아, 딕셔너리를 불러 필요한 정보를 가져옴 
## 작품번호, 제목, 작자, 창작연대, 검색된 행 등을 Excel로 저장하여 제공토록 함
# 추출함수로 만들어 코드를 간략하게 함.


import os, re
import glob, time
import sys, codecs
import 말뭉치_함수RE
import Sizo_Dic
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


global fff, words, i, count_item, grab, startend, howmany, grabb, each_line1

# ㄷㅏㄱ?ㅏ> .{3,25}ㄴ.ㄴㄷㅡㅅ" # 성공 "ㄴᆞㄴ\s\S.{1,15}ㅇㅏ\s"    /ㅇㅔㅅㅕ.{1,9}ㄹ\s?ㅅㅗㄴㅑ/
#words = r"\b\w*?(\w{4,6})\b.{1,3}?\b\w*?\1\b"
words = "ㄷㅗㄷㅗ[ㅡ|ᆞ]ㄴㄷㅏ.{3,30}[ㅏ|ㅑ]"
belong = "송성42중_"
words_han ="_는다__아(야)"
# 정규표현식 문법 ?(앞 문자가 0개 또는 1개)/ .(최소 한개의 문자/공백포함)/ {n,m}(n번 이상 m번 이하 반복: dot와 조합시켜야)/ \s(스페이스) / \S(스페이스 아닌 거) / (그룹){반복횟수}
# 한자로 된 작품은 한음절 크기가 1임에 주의할 것. {범위} 입력 시 1부터 입력할 것.
# \s -> 그냥 띄어쓰기도 됨  #아래아 + ㅣ = 한글자(ᆡ) / .{1,2} 이런 형태로 검색해야 함.
# 띄어쓰기 유무에 주의 -> 선택적으로 검색 : \s?
### 의 다르니  ㅇㅡㅣ ㄴᆡㄷㅏㄹㅡㄴㅣ 
# 연속된 같은 문자 검색 ([a-z])\1
# "(^ㅇㅣ.{1,15}ㄴ\s.{2,40}\sㅈㅕ.{1,15}ㄴ\s)|(\sㅇㅣ.{1,15}ㄴ\s.{2,40}\sㅈㅕ.{1,15}ㄴ\s)" # 성공(그룹으로 선택)
# "\S.{4,11}ㅇㅣ\s\S.{4,11}ㄱㅗ\S.{4,11}ㅇㅣ\s\S.{4,11}ᆡ" # 성공
# "ㄴᆞㄴ\s\S{0,1}ㅈㅓ\s\S.{1,8}ㅇㅏ\s" # 성공
#"ㄴᆞㄴ\s\S{0,1}[ㅈㅓ |ㄷㅕ |ㅈㅕ ]\s\S.{1,8}[ㅇㅏ |ㅇㅑ ]\s" # 성공
# "冬+.{0,6}ㅅ+.{2,4}[ㄷ|ㅼ]+.{1,1}ㄹ"     |ㄷㅕ |ㅈㅕ  |ㅇㅣㅇㅕ |ㅇㅑ 
# "ㄴᆞㄴ\s\S{1,5}ㄹ\s\S.{1,15}ㄴᆞㄴ\s\S{1,5}ㄹ\s" # 는 ㄹ 는 ㄹ #성공

mal="sizo_work"  
#"_data_history"  
#fff=""
fff=[]
sub_fff=[]
startend=[]
grab=1  # 홀수이면 선행 행의 값을 반영하지 않음 
howmany=0
grabb=0
each_line1=""

def sub_fff_TO_fff(count_number,w_num,genre1,author1,author_year, book1, book1_year, each_line, line1, line2, line3, copynum, fff_list):
    count_number+=1
    each_line = each_line.strip('\r')
    sub_fff=[]
    sub_fff.append(count_number)
    sub_fff.append(w_num)
    sub_fff.append(genre1)
    sub_fff.append(author1)
    sub_fff.append(author_year)
    sub_fff.append(book1)
    sub_fff.append(book1_year)
    sub_fff.append(each_line)
    sub_fff.append(line1)
    sub_fff.append(line2)
    sub_fff.append(line3)
    sub_fff.append(copynum)
    fff_list.append(sub_fff)
    return count_number, fff_list

                                    


#path="gasa_work_paired" #아래처럼 절대경로로 바꾸어주니 작동함
#path="C:\시조_수사_전승\sizo_scan\sizo_work_paired"
path="C:\\Users\\park9\\OneDrive\\바탕 화면\\시조_수사_전승\\sizo_scan\\sizo_work_paired"

data_files = glob.glob(f"{path}/*") #

count_item=0

errors = ""

try:
    for e_f in data_files:
        i=0
        
        with codecs.open(e_f,"r", encoding="utf-16") as f:
            name=f.readline()
            name=f.readline()
            #print(name0)
            print(name)
            #time.sleep(1)
            
            # 작품번호, 작품제목 추출  --> 시조는 제목이 없음/ 어떡하나???
            try:
                #work_num, work_name = name.split(".")
                work_num = int(name[1:])
                work_name = "00000"
            except Exception as ex: # 에러 종류
                print("에러가 발생했습니다.", ex)
                errors = errors + name +"\n"
                #print(name)
 
    
            work_num = int(work_num)
            
            
            try:
                for each_line in f:
                    pattern_search = re.compile(words)
                    c = pattern_search.findall(each_line)  # 검색 건수를 아래에서 howmany로 저장할 것.

                    if  (grab%2) == 0:
                        each_line1=each_line
                    
                   
                        
                    
                    if len(c)==0 and (grab%2) != 0  : #검색 값이 없을 때, 선행 행의 값을 가져오지 않음  
                        pass
                    elif len(c)>0 and (grab%2) != 0 : #자모열에서 검색 값이 있을 때, 선행 행의 값을 가져오지 않은 상태,  AAAAAA로 넘겨
                        grab+=1 #선행 행의 값을 다음 행에 연결하도록 짝수로 만듦
                                           
                        
                        for m in re.finditer(words,each_line): #검색 구간을 검색하여 비율값을 리스트로 저장
                            start=int(m.start())
                            end=int(m.end())
                            startend=[]
                            
                            length=len(each_line) # 문장의 길이

                            pc_start=(start/length) #시작점의 비율
                            pc_end=(end/length) #끝점의 비율

                            
                            startend.append(pc_start)
                            startend.append(pc_end)

                            howmany=len(c)
                            
                            
                     
                    elif  (grab%2) == 0: #AAAAAAA 자모열의 검색 값을 넘겨 받아서 문자열의 검색 값을 저장 및 출력  grab은 짝수상태(검색유)
                        
                        grab+=1  # 다시 홀수로 만들어 닫음. 

                        

                        #print("AAA111111111111111111111111111111111111111AAAA")
                        #print(each_line1)
                        
                        
                        
                    
                        if i==0: # 문서의 처음/Sizo_Dic과 연동하여 '작가명/창작연대' 변수로 가져올 것
                            #딕셔너리와 연동 -> 작가/창작연대 추출 -> 아래에 이 변수를 반복 활용할 것.
                            #새 화일의 첫줄에서 자동 갱신됨.
                            print('work_num :' + str(work_num))
                            #print('dic count :' + str(len(Gasa_Dic.gasa_dic['가사'])))
                            #print('1004번 작품 정보 :' + str(Gasa_Dic.gasa_dic['가사'][1004]))
                            
                            
                            #'이본수': 'copy_num', '문헌명': 'book', '문헌연대':'book_year', '장르':'genre', 
                            #'작가명':'writer', '작가연대':'writer_year', '초장':'line01', '중장':'line02', '종장':'line03' 
                            
                            copy_num = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['이본수']
                            book = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['문헌명']
                            book_year = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['문헌연대']
                            genre = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['장르']
                            author = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['작가명']
                            author_year = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['작가연대']                            
                            line01 = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['초장']
                            line02 = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['중장']
                            line03 = Sizo_Dic.sizo_dic['교본역대시조전서'][work_num]['종장']

                            #author.pop('\n')
                            
                            sub_fff=[]
                            #fff=fff+"\n"+"\n"+"\n"+"\n"+"\n"+name+"\n"
                            i=1
                            lenlen=int(len(startend)/2)
                            print('len(each_line1)                        :' +str(len(each_line1))) #확인용>>>>>>>>>>>>>>>>>>

                            le=len(each_line1) #문장 길이
                            lc=howmany     #조회된 갯수   #len(each_line[new_start:new_end])
                            lw=(len(words)/5)

                            if 0<lc<2:
                                ll=int((le-lw)/(lc+1)) #조회할 앞 뒤 단어의 길이 계산
                                if ll>30:
                                    ll=32

                            elif 1<lc:
                                ll=int((le-lw)/lc) #조회할 앞 뒤 단어의 길이 계산
                                    
                                if ll>30:
                                    ll=32
                            else:
                                pass

                                
                            for m in range(lenlen):
                                
                                ppc_start=startend.pop(0)
                                ppc_end=startend.pop(0)
                                new_start=int(len(each_line1)*ppc_start)
                                new_end=int(len(each_line1)*ppc_end)
                                
                                
                                if len(each_line1)<50:
                                    each_line_chosen = each_line1[:]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff) 
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[:]+"\n"+"\n"
                                    #print('1111')
                                    '''print('le:' + str(le))
                                    print('len(each_line):' +str(len(each_line1)))'''
                                    
                                elif ll>new_start and (le-new_end)>ll: 
                                    each_line_chosen = each_line1[:new_end+ll]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[:new_end+ll]+"\n"+"\n"
                                                                       
                                    #print('1')
                            
                                elif ll<new_start and (le-new_end)>ll:
                                    each_line_chosen = each_line1[new_start-ll:new_end+ll]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[new_start-ll:new_end+ll]+"\n"+"\n"
                                    #print('2')
                                    
                                          
                            
                                elif ll<new_start and (le-new_end)<ll:
                                    each_line_chosen = each_line1[new_start-ll:]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[new_start-ll:]+"\n"+"\n"
                
                            
                                elif ll>new_start and (le-new_end)<ll:
                                    each_line_chosen = each_line1[:]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[:]+"\n"+"\n"
                                    
                                    #print('4')
                                else:
                                    each_line_chosen = each_line1[new_start:new_end]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+str(count_item)+")"+ each_line1[new_start:new_end]+"\n"+"\n"
                                    #print('5')
                                    
                        
                            
                            
                            
                        elif i>0:
                            
                            lenlen=int(len(startend)/2)
                            
                            for m in range(lenlen):
                                ppc_start=startend.pop(0)
                                ppc_end=startend.pop(0)
                                new_start=int(len(each_line1)*ppc_start)
                                new_end=int(len(each_line1)*ppc_end)
                                
                                le=len(each_line1) #문장 길이
                                lc=len(each_line1[new_start:new_end])  #조회된 갯수
                                lw=(len(words)/5)

                                if 0<lc<2:
                                    ll=int((le-lw)/(lc+1)) #조회할 앞 뒤 단어의 길이 계산

                                elif 1<lc:
                                    ll=int((le-lw)/lc) #조회할 앞 뒤 단어의 길이 계산
                                    
                                    if ll>30:
                                        ll=22
                                else:
                                    pass


                                if len(each_line)<50: 
                                    each_line_chosen = each_line1[:]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[:]+"\n"+"\n"
                                    #print('2222')
                                    #print(each_line1)
                                    
                                elif ll>new_start and (le-new_end)>ll:
                                    each_line_chosen = each_line1[:new_end+ll]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[:new_end+ll]+"\n"+"\n"
                                    #print('6')
                            
                                elif ll<new_start and (le-new_end)>ll:
                                    each_line_chosen = each_line1[new_start-ll:new_end+ll]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[new_start-ll:new_end+ll]+"\n"+"\n"
                                
                            
                                elif ll<new_start and (le-new_end)<ll:
                                    each_line_chosen = each_line1[new_start-ll:]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[new_start-ll:]+"\n"+"\n"
                                    #print('8')
                            
                                elif ll>new_start and (le-new_end)<ll:
                                    each_line_chosen = each_line1[:]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+ str(count_item)+")"+each_line1[:]+"\n"+"\n"
                                    #print('9')

                                    
                                else:
                                    each_line_chosen = each_line1[new_start:new_end]
                                    count_item, fff = sub_fff_TO_fff(count_item,work_num,genre,author,author_year, book, book_year, each_line, line01, line02, line03, copy_num, fff)
                                    #count_item+=1
                                    #fff=fff+str(count_item)+")"+ each_line1[new_start:new_end]+"\n"+"\n"
                                    #print('10')
                        
                                
                             
                    elif each_line==f.readline(-1):
                        f.close()
                        print("f.close()니다")
                        
                    else:
                        
                        pass
                    
                    
            except IOError as ioerr:
                print('File error (outer try): ' + str(ioerr))
    
                
    if not len(fff)==0 :
        #out=open('C:\\가사작업중20200208\\Gasa_Divide\\result\\result_'+'('+str(count_item)+'건)'+words_han+"_from_"+mal+'.txt', mode= 'w', encoding='utf-16')
        #print(fff, file=out)
        
        #엑셀로 저장
        wb = Workbook()
        ws = wb.active
        ws.title = words_han
        thin_border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
        
        x = 1
        y = 0 
        for s_fff in fff:
            #제목 행 세팅
            ws.cell(column = 1, row = 1, value = "연번").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 1, row = 1, value = "연번").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 2, row = 1, value = "작품번호").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 2, row = 1, value = "작품번호").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 3, row = 1, value = "장르").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 3, row = 1, value = "장르").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 4, row = 1, value = "작가").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 4, row = 1, value = "작가").font = Font(name="맑은 고딕", size=11, color="00FF0000")
            ws.cell(column = 5, row = 1, value = "작가연대").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 5, row = 1, value = "작가연대").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 6, row = 1, value = "가집").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 6, row = 1, value = "가집").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 7, row = 1, value = "가집연대").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 7, row = 1, value = "가집연대").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 8, row = 1, value = "패턴발견구절").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 8, row = 1, value = "패턴발견구절").font = Font(name="맑은 고딕", size=10, color="00FF0000")
            ws.cell(column = 9, row = 1, value = "초장").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 9, row = 1, value = "초장").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 10, row = 1, value = "중장").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 10, row = 1, value = "중장").font = Font(name="맑은 고딕", size=11, color="00FF0000")
            ws.cell(column = 11, row = 1, value = "종장").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 11, row = 1, value = "종장").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            ws.cell(column = 12, row = 1, value = "이본수").border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
            ws.cell(column = 12, row = 1, value = "이본수").font = Font(name="맑은 고딕", size=12, color="00FF0000")
            


       

            # 2열부터 검색된 내용 정리
            x+=1
            y=0
            for item in s_fff:
                y+=1
                #ws.cell = Font(name ="NGULIM", size = 20)
                cell_address = ws.cell(column = y, row = x, value = item)
                #print(cell_address.coordinate)
                cell_address_obj = ws[cell_address.coordinate]
                cell_address_obj.font = Font(name ="새굴림", size = 10)
                
                #ws[aaaa].font  = Font(name ="NGULIM", size = 20)
                #ws.cell = Font(name ="NGULIM", size = 20)
                cell_address_obj.border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
        
        # 열의 넓이 정하기
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 7
        ws.column_dimensions['C'].width = 4
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 60
        ws.column_dimensions['I'].width = 60
        ws.column_dimensions['J'].width = 60
        ws.column_dimensions['K'].width = 60
        ws.column_dimensions['L'].width = 6
                
        wb.save('C:\\Users\\park9\\OneDrive\\바탕 화면\\시조_수사_전승\\sizo_scan\\result\\'+belong+'result_'+'('+str(count_item)+'건)'+words_han+"_from_"+mal+'.xlsx')
        wb.close
       
        
        #ffff="error checking"   # 저장 경로 확인해 보니 c:\사용자\park9 
        #with open('resul9999t.txt', mode= 'w', encoding= 'utf-16') as f:
        #    f.write(fff)
                  
        #out.close()
        print("검색어는 '" +str(words)+"' 입니다")
        
        print("검색 건수는 모두 '" +str(count_item)+"'건 입니다. 엑셀 파일로 저장하였습니다. ")
        
        
    else:
        print("해당되는 자료가 없습니다")
        
                    
except IOError as ioerr:
    print('File error (outer try): ' + str(ioerr))

print(errors)
    


    
    
    
